from __future__ import annotations
import os
import re
import io
import uuid
import json
import shutil
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any
from zipfile import ZipFile
from lxml import etree
from PIL import Image
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation

def _json_default(obj: Any):
    """
    JSON-serializator for non-standard types inside index.
    Сейчас нам нужен прежде всего для Decimal в chart-структурах.
    """
    if isinstance(obj, Decimal):
        # храним числовое значение как float
        return float(obj)
    # для любых других типов пусть json выбрасывает нормальную ошибку
    raise TypeError(f"Object of type {obj.__class__.__name__} is not JSON serializable")

NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "pic": "http://schemas.openxmlformats.org/drawingml/2006/picture",
}

FIG_CAP_RE = re.compile(
    r"^(?:Рис\.?|Рисунок|Figure)\s*\.?\s*(?:№\s*)?([A-Za-zА-Яа-я]?\s*\d+(?:[.,]\d+)*)\b.*",
    re.IGNORECASE,
)
TAB_CAP_RE = re.compile(
    r"^(?:Таблица|Table)\s*\.?\s*(?:№\s*)?([A-Za-zА-Яа-я]?\s*\d+(?:[.,]\d+)*)\b.*",
    re.IGNORECASE,
)

HEAD_REF_RE = re.compile(
    r"^(Список\s+(?:использованных\s+)?(?:литературы|источников)|Библиографический\s+список|References)$",
    re.IGNORECASE,
)
HEAD_STYLE_RE = re.compile(r"(heading|заголовок)\s*([1-6])", re.IGNORECASE)

RUNTIME_DIR = os.getenv("RUNTIME_DIR", "runtime")
MEDIA_ROOT = os.path.join(RUNTIME_DIR, "media")
INDEX_ROOT = os.path.join(RUNTIME_DIR, "indexes")

MAX_TABLE_ROWS = int(os.getenv("DOC_MAX_TABLE_ROWS", "50"))
MAX_SERIES_POINTS = int(os.getenv("DOC_MAX_SERIES_POINTS", "200"))


@dataclass
class Heading:
    level: int
    text: str


@dataclass
class Figure:
    kind: str
    caption: Optional[str] = None
    n: Optional[int] = None
    image_path: Optional[str] = None
    chart: Optional[Dict] = None
    _order: int = field(default=0)
    # индекс параграфа, в котором встретили этот drawing (для привязки к ближайшей подписи)
    para_idx: int = field(default=-1)



@dataclass
class Table:
    rows: List[List[str]]
    caption: Optional[str] = None
    n: Optional[int] = None
    _order: int = field(default=0)
    # индекс параграфа, рядом с которым встретилась таблица
    para_idx: int = field(default=-1)


def _read_xml(z: ZipFile, path: str) -> etree._Element:
    data = z.read(path)
    return etree.fromstring(data)


def _text_of(node: etree._Element) -> str:
    parts = node.xpath(".//w:t/text()", namespaces=NS)
    return "".join(parts).replace("\xa0", " ").strip()


def _doc_rels(z: ZipFile, rels_path: str) -> Dict[str, Dict[str, str]]:
    rels: Dict[str, Dict[str, str]] = {}
    if rels_path in z.namelist():
        root = _read_xml(z, rels_path)
        for rnode in root.findall(".//", namespaces=NS):
            if isinstance(rnode.tag, str) and rnode.tag.endswith("Relationship"):
                rid = rnode.get(f"{{{NS['r']}}}Id") or rnode.get("Id")
                tgt = rnode.get("Target")
                typ = rnode.get("Type")
                if rid and tgt:
                    rels[rid] = {"Target": tgt, "Type": typ or ""}
    return rels


def _ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)


def _iter_body_children(doc: etree._Element):
    body = doc.find("w:body", namespaces=NS)
    if body is None:
        return
    for child in body:
        if isinstance(child.tag, str) and (
            child.tag.endswith("p") or child.tag.endswith("tbl")
        ):
            yield child


def _extract_heading_level(p: etree._Element) -> Optional[int]:
    val = p.xpath("w:pPr/w:pStyle/@w:val", namespaces=NS)
    if not val:
        return None
    sid = (val[0] or "").strip()
    m = HEAD_STYLE_RE.search(sid)
    if m:
        try:
            return int(m.group(2))
        except Exception:
            return None
    return None



def _parse_tbl(tbl: etree._Element, limit_rows: int = MAX_TABLE_ROWS) -> List[List[str]]:
    rows: List[List[str]] = []
    for tr in tbl.findall("w:tr", namespaces=NS):
        row: List[str] = []
        for tc in tr.findall("w:tc", namespaces=NS):
            txt = " ".join(
                _text_of(p) for p in tc.findall(".//w:p", namespaces=NS) or []
            )
            row.append(txt.strip())
        if any(cell for cell in row):
            rows.append(row)
        if len(rows) >= limit_rows:
            break
    return rows


def _save_image_from_rel(z: ZipFile, rel_target: str, doc_id: str) -> Optional[str]:
    internal = "word/" + rel_target.lstrip("/\\")
    if internal not in z.namelist():
        internal = os.path.normpath(os.path.join("word", rel_target)).replace("\\", "/")
        if internal not in z.namelist():
            return None
    data = z.read(internal)
    ext = os.path.splitext(internal)[1].lower() or ".bin"
    out_dir = os.path.join(MEDIA_ROOT, doc_id)
    _ensure_dir(out_dir)
    fname = f"img_{uuid.uuid4().hex[:8]}{ext}"
    out_path = os.path.join(out_dir, fname)
    with open(out_path, "wb") as f:
        f.write(data)
    try:
        with Image.open(out_path) as im:
            im.verify()
    except Exception:
        pass
    return out_path


def _fmt_percent_value(v: float) -> str:
    p = v * 100.0
    if abs(p - round(p)) < 0.05:
        return f"{int(round(p))}%"
    s = f"{p:.2f}".rstrip("0").rstrip(".")
    return f"{s}%"


def _has_percent_numfmt(root: etree._Element) -> bool:
    for nm in root.findall(".//c:numFmt", namespaces=NS):
        fc = (nm.get("formatCode") or "").lower()
        if "%" in fc:
            return True
    return False


def _is_percent_grouping(root: etree._Element) -> bool:
    for g in root.findall(".//c:grouping", namespaces=NS):
        val = (g.get("val") or g.get(f"{{{NS['c']}}}val") or "").lower()
        if "percentstacked" in val:
            return True
    return False


def _vals_flat(series: List[Dict]) -> List[float]:
    out: List[float] = []
    for s in series:
        for v in (s.get("vals") or []):
            try:
                out.append(float(v))
            except Exception:
                pass
    return out


def _sum_close_to_one_for_pie(series: List[Dict]) -> bool:
    if not series:
        return False
    raw = series[0].get("vals") or []
    vals: List[float] = []
    for v in raw:
        try:
            vals.append(float(v))
        except Exception:
            pass
    if not vals:
        return False
    s = sum(vals)
    return (0.98 <= s <= 1.02) and max(vals) <= 1.0


def _per_category_sums_close_to_one(series: List[Dict]) -> bool:
    if not series:
        return False
    max_len = max(len(s.get("vals") or []) for s in series)
    if max_len == 0:
        return False
    sums: List[float] = []
    for i in range(max_len):
        acc = 0.0
        present = False
        for s in series:
            vals = s.get("vals") or []
            if i < len(vals):
                try:
                    acc += float(vals[i])
                    present = True
                except Exception:
                    pass
        if present:
            sums.append(acc)
    if len(sums) < 2:
        return False
    ok = [0.95 <= x <= 1.05 for x in sums]
    return sum(1 for x in ok if x) >= 2


def _dec_from_any(x: Any) -> Optional[Decimal]:
    if x is None:
        return None
    if isinstance(x, (int, Decimal)):
        return Decimal(x)
    if isinstance(x, float):
        return Decimal(str(x))
    try:
        s = str(x).strip().replace("\u00A0", " ").replace(" ", "")
        s = s.replace(",", ".")
        if s.endswith("%"):
            s = s[:-1]
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _dec_list_to_str(lst: List[Optional[Decimal]]) -> List[str]:
    out: List[str] = []
    for d in lst:
        if d is None:
            continue
        s = format(d.normalize(), "f").rstrip("0").rstrip(".")
        out.append(s if s != "" else "0")
    return out


def _trim_pair_lists(a: List, b: List) -> Tuple[List, List]:
    n = min(len(a), len(b))
    return a[:n], b[:n]


def _read_cache_texts(parent: etree._Element, path: str) -> List[str]:
    pts = parent.findall(path, namespaces=NS)
    out: List[str] = []
    for p in pts[:MAX_SERIES_POINTS]:
        val = (p.text or "").strip()
        out.append(val)
    return out


def _read_cache_nums(parent: etree._Element, path: str) -> List[str]:
    """
    Старое поведение: возвращает нормализованные числовые строки
    (без информации о формате). Нужен для scatter/bubble и части эвристик.
    """
    pts = parent.findall(path, namespaces=NS)
    out: List[str] = []
    for p in pts[:MAX_SERIES_POINTS]:
        txt = (p.text or "").strip()
        if not txt:
            continue
        d = _dec_from_any(txt)
        if d is not None:
            s = format(d.normalize(), "f").rstrip("0").rstrip(".")
            out.append(s if s != "" else "0")
    return out


def _read_cache_nums_with_raw(parent: etree._Element, path: str) -> List[dict]:
    """
    Читаем numCache так, чтобы учитывать индекс точки (pt/@idx).

    В OOXML нулевые значения часто просто пропускаются в кэше:
      <c:pt idx="1">70</c:pt>
      <c:pt idx="2">80</c:pt>
    и т.п.

    Нам нужно вернуть плотный список 0..max_idx, заполняя пропуски
    {"raw": "", "dec": None}, чтобы не съезжало выравнивание
    «категория ↔ значение».
    """
    # Игнорируем `path` и работаем именно с <c:pt>
    pts = parent.findall(".//c:numCache//c:pt", namespaces=NS)
    if not pts:
        pts = parent.findall(".//c:numLit//c:pt", namespaces=NS)
    if not pts:
        return []

    items: List[dict] = []
    max_idx: int = -1

    # если у последнего <c:pt> нет атрибута idx — добавляем вручную
    for i, p in enumerate(pts):
        if p.get("idx") is None:
            p.set("idx", str(i))


    for p in pts[:MAX_SERIES_POINTS]:
        idx_raw = p.get("idx")
        try:
            idx = int(idx_raw) if idx_raw is not None else None
        except Exception:
            idx = None

        v_node = p.find("c:v", namespaces=NS)
        raw = (v_node.text or "").strip() if v_node is not None and v_node.text else ""
        dec = _dec_from_any(raw) if raw else None

        items.append({"idx": idx, "raw": raw, "dec": dec})
        if idx is not None and idx > max_idx:
            max_idx = idx

    # Если нет корректных idx — возвращаем как раньше (просто список по порядку)
    if max_idx < 0:
        return [{"raw": it["raw"], "dec": it["dec"]} for it in items]

    # Строим плотный список 0..max_idx, подставляя пустые слоты
    dense: List[dict] = [{"raw": "", "dec": None} for _ in range(max_idx + 1)]
    for it in items:
        idx = it["idx"]
        if idx is None or idx < 0 or idx > max_idx:
            continue
        dense[idx] = {"raw": it["raw"], "dec": it["dec"]}

    if len(dense) < len(items):
    # если длина меньше — расширяем
        while len(dense) < len(items):
            dense.append({"raw": "", "dec": None})

    return dense


def _parse_scatter_series(ser: etree._Element) -> Dict:
    name_node = ser.find(".//c:tx//c:v", namespaces=NS) or ser.find(
        ".//c:tx//c:rich//a:t", namespaces=NS
    )
    sname = (
        name_node.text.strip()
        if (name_node is not None and name_node.text)
        else None
    )
    x_parent = ser.find(".//c:xVal", namespaces=NS)
    y_parent = ser.find(".//c:yVal", namespaces=NS)
    x = (
        _read_cache_nums(x_parent, ".//c:numCache//c:pt//c:v")
        if x_parent is not None
        else []
    )
    if not x and x_parent is not None:
        x = _read_cache_nums(x_parent, ".//c:numLit//c:pt//c:v")
    y = (
        _read_cache_nums(y_parent, ".//c:numCache//c:pt//c:v")
        if y_parent is not None
        else []
    )
    if not y and y_parent is not None:
        y = _read_cache_nums(y_parent, ".//c:numLit//c:pt//c:v")
    x, y = _trim_pair_lists(x, y)
    return {"name": sname, "x_vals_num": x, "y_vals_num": y}


def _parse_bubble_series(ser: etree._Element) -> Dict:
    name_node = ser.find(".//c:tx//c:v", namespaces=NS) or ser.find(
        ".//c:tx//c:rich//a:t", namespaces=NS
    )
    sname = (
        name_node.text.strip()
        if (name_node is not None and name_node.text)
        else None
    )
    x_parent = ser.find(".//c:xVal", namespaces=NS)
    y_parent = ser.find(".//c:yVal", namespaces=NS)
    size_parent = ser.find(".//c:bubbleSize", namespaces=NS)
    x = (
        _read_cache_nums(x_parent, ".//c:numCache//c:pt//c:v")
        if x_parent is not None
        else []
    )
    if not x and x_parent is not None:
        x = _read_cache_nums(x_parent, ".//c:numLit//c:pt//c:v")
    y = (
        _read_cache_nums(y_parent, ".//c:numCache//c:pt//c:v")
        if y_parent is not None
        else []
    )
    if not y and y_parent is not None:
        y = _read_cache_nums(y_parent, ".//c:numLit//c:pt//c:v")
    b = (
        _read_cache_nums(size_parent, ".//c:numCache//c:pt//c:v")
        if size_parent is not None
        else []
    )
    if not b and size_parent is not None:
        b = _read_cache_nums(size_parent, ".//c:numLit//c:pt//c:v")
    n = min(len(x), len(y), len(b))
    return {
        "name": sname,
        "x_vals_num": x[:n],
        "y_vals_num": y[:n],
        "size_vals_num": b[:n],
    }


def _parse_chart_data(z: ZipFile, chart_xml_path: str) -> Dict:
    root = _read_xml(z, chart_xml_path)
    plot = root.find(".//c:plotArea", namespaces=NS)
    chart_type: Optional[str] = None
    chart_node: Optional[etree._Element] = None

    if plot is not None:
        for tname in (
            "barChart",
            "lineChart",
            "pieChart",
            "areaChart",
            "scatterChart",
            "radarChart",
            "bar3DChart",
            "ofPieChart",
            "doughnutChart",
            "colChart",
            "col3DChart",
            "bubbleChart",
        ):
            node = plot.find(f"c:{tname}", namespaces=NS)
            if node is not None:
                chart_type = tname.replace("Chart", "").lower()
                chart_node = node
                break

    def _axis_title(ax_xpath: str) -> Optional[str]:
        title = root.find(ax_xpath + "/c:title", namespaces=NS)
        if title is None:
            return None
        t = title.find(".//c:tx//c:rich//a:t", namespaces=NS)
        if t is not None and t.text:
            return t.text.strip()
        v = title.find(".//c:tx//c:v", namespaces=NS)
        if v is not None and v.text:
            return v.text.strip()
        return None

    x_title = _axis_title(".//c:catAx") or _axis_title(".//c:dateAx")
    y_title = _axis_title(".//c:valAx") or _axis_title(".//c:serAx")

    title_node = root.find(".//c:title", namespaces=NS)
    chart_title: Optional[str] = None
    if title_node is not None:
        tt = title_node.find(".//c:tx//c:rich//a:t", namespaces=NS) or title_node.find(
            ".//c:tx//c:v", namespaces=NS
        )
        if tt is not None and tt.text:
            chart_title = tt.text.strip()

    series: List[Dict[str, Any]] = []
    used_source = "cache"
    percent_hint_from_xlsx = False

    # 1. Чтение данных из кэша OOXML
    if chart_node is not None and chart_type in ("scatter", "bubble"):
        for ser in chart_node.findall("c:ser", namespaces=NS):
            if chart_type == "scatter":
                series.append(_parse_scatter_series(ser))
            else:
                series.append(_parse_bubble_series(ser))
    elif chart_node is not None:
        # категориальные диаграммы: bar, line, pie, area, col, ...
        for ser in chart_node.findall("c:ser", namespaces=NS):
            name = ser.find(".//c:tx//c:v", namespaces=NS)
            if name is None:
                name = ser.find(".//c:tx//c:rich//a:t", namespaces=NS)
            series_name = (
                name.text.strip()
                if (name is not None and name.text)
                else None
            )

            cats: List[str] = []
            cat_parent = ser.find(".//c:cat", namespaces=NS)
            if cat_parent is not None:
                cats = _read_cache_texts(cat_parent, ".//c:strCache//c:pt//c:v")
                if not cats:
                    cats = _read_cache_nums(
                        cat_parent, ".//c:numCache//c:pt//c:v"
                    )

            values_raw_dec: List[dict] = []
            val_parent = ser.find(".//c:val", namespaces=NS)
            if val_parent is not None:
                values_raw_dec = _read_cache_nums_with_raw(
                    val_parent, ".//c:numCache//c:pt//c:v"
                )
                if not values_raw_dec:
                    values_raw_dec = _read_cache_nums_with_raw(
                        val_parent, ".//c:numLit//c:pt//c:v"
                    )

            if cats and values_raw_dec:
                n = min(len(cats), len(values_raw_dec))
                cats = cats[:n]
                values_raw_dec = values_raw_dec[:n]

            series.append(
                {
                    "name": series_name,
                    "cats": cats,
                    "values_raw_dec": values_raw_dec,
                }
            )

    # 2. Если данных нет или они неполные – пробуем xlsx-источник
    needs_xlsx = False
    if chart_type in ("scatter", "bubble"):
        if not series or any(
            (len(s.get("x_vals_num", [])) == 0 or len(s.get("y_vals_num", [])) == 0)
            for s in series
        ):
            needs_xlsx = True
    else:
        needs_xlsx = any(
            (not s.get("cats") or not s.get("values_raw_dec"))
            for s in series
        ) or (not series)

    if needs_xlsx:
        base_dir = os.path.dirname(chart_xml_path)
        rels_path = os.path.join(
            base_dir, "_rels", os.path.basename(chart_xml_path) + ".rels"
        ).replace("\\", "/")
        rels = _doc_rels(z, rels_path)
        xlsx_target = None
        for r in rels.values():
            tgt = r.get("Target", "")
            if "embeddings/" in tgt and tgt.lower().endswith(".xlsx"):
                xlsx_target = tgt
                break
        if xlsx_target:
            used_source = "xlsx"
            xlsx_internal = os.path.normpath(
                os.path.join(base_dir, xlsx_target)
            ).replace("\\", "/")
            if xlsx_internal not in z.namelist():
                xlsx_internal = os.path.normpath(
                    os.path.join("word", xlsx_target)
                ).replace("\\", "/")
            if xlsx_internal in z.namelist():
                wb = load_workbook(
                    filename=io.BytesIO(z.read(xlsx_internal)), data_only=True
                )

                def read_range_plus(fstr: str) -> Tuple[List, bool]:
                    if not fstr or "!" not in fstr:
                        return [], False
                    sheet, rng = fstr.split("!", 1)
                    sheet = sheet.strip().strip("'").strip('"')
                    if sheet not in wb.sheetnames:
                        return [], False
                    ws = wb[sheet]
                    rng = rng.replace("$", "")
                    if ":" in rng:
                        a, b = rng.split(":", 1)
                    else:
                        a = b = rng
                    try:
                        cells = ws[a:b]
                    except Exception:
                        return [], False
                    out: List[Any] = []
                    percent_fmt = False
                    if isinstance(cells, tuple):
                        for row in cells:
                            for c in (
                                row if isinstance(row, tuple) else (row,)
                            ):
                                out.append(c.value)
                                try:
                                    nf = str(
                                        getattr(c, "number_format", "") or ""
                                    ).lower()
                                    if "%" in nf:
                                        percent_fmt = True
                                except Exception:
                                    pass
                    else:
                        out.append(cells.value)
                        try:
                            nf = str(
                                getattr(cells, "number_format", "") or ""
                            ).lower()
                            if "%" in nf:
                                percent_fmt = True
                        except Exception:
                            pass
                    return out, percent_fmt

                series_tmp: List[Dict[str, Any]] = []

                if chart_node is not None and chart_type in (
                    "scatter",
                    "bubble",
                ):
                    # scatter / bubble через xlsx
                    for ser in chart_node.findall("c:ser", namespaces=NS):
                        nm = ser.find(
                            ".//c:tx//c:v", namespaces=NS
                        ) or ser.find(".//c:tx//c:rich//a:t", namespaces=NS)
                        series_name = (
                            nm.text.strip()
                            if (nm is not None and nm.text)
                            else None
                        )

                        x_ref = ser.find(
                            ".//c:xVal//c:numRef//c:f", namespaces=NS
                        )
                        y_ref = ser.find(
                            ".//c:yVal//c:numRef//c:f", namespaces=NS
                        )
                        x_vals, x_pct = read_range_plus(
                            x_ref.text if x_ref is not None and x_ref.text else ""
                        )
                        y_vals, y_pct = read_range_plus(
                            y_ref.text if y_ref is not None and y_ref.text else ""
                        )
                        percent_hint_from_xlsx = (
                            percent_hint_from_xlsx or x_pct or y_pct
                        )
                        x_num = _dec_list_to_str(
                            [_dec_from_any(v) for v in x_vals[:MAX_SERIES_POINTS]]
                        )
                        y_num = _dec_list_to_str(
                            [_dec_from_any(v) for v in y_vals[:MAX_SERIES_POINTS]]
                        )
                        x_num, y_num = _trim_pair_lists(x_num, y_num)
                        if chart_type == "bubble":
                            b_ref = ser.find(
                                ".//c:bubbleSize//c:numRef//c:f", namespaces=NS
                            )
                            b_vals, b_pct = read_range_plus(
                                b_ref.text
                                if b_ref is not None and b_ref.text
                                else ""
                            )
                            percent_hint_from_xlsx = (
                                percent_hint_from_xlsx or b_pct
                            )
                            b_num = _dec_list_to_str(
                                [
                                    _dec_from_any(v)
                                    for v in b_vals[:MAX_SERIES_POINTS]
                                ]
                            )
                            n = min(len(x_num), len(y_num), len(b_num))
                            series_tmp.append(
                                {
                                    "name": series_name,
                                    "x_vals_num": x_num[:n],
                                    "y_vals_num": y_num[:n],
                                    "size_vals_num": b_num[:n],
                                }
                            )
                        else:
                            series_tmp.append(
                                {
                                    "name": series_name,
                                    "x_vals_num": x_num,
                                    "y_vals_num": y_num,
                                }
                            )
                elif chart_node is not None:
                    # категориальные диаграммы через xlsx
                    for ser in chart_node.findall("c:ser", namespaces=NS):
                        nm = ser.find(
                            ".//c:tx//c:v", namespaces=NS
                        ) or ser.find(".//c:tx//c:rich//a:t", namespaces=NS)
                        series_name = (
                            nm.text.strip()
                            if (nm is not None and nm.text)
                            else None
                        )

                        cats: List[str] = []
                        cat_ref = ser.find(
                            ".//c:cat//c:strRef//c:f", namespaces=NS
                        )
                        if cat_ref is None:
                            cat_ref = ser.find(
                                ".//c:cat//c:numRef//c:f", namespaces=NS
                            )
                        if cat_ref is not None and cat_ref.text:
                            cat_vals, cat_pct = read_range_plus(cat_ref.text)
                            cats = [
                                ("" if v is None else str(v))
                                for v in cat_vals[:MAX_SERIES_POINTS]
                            ]
                            percent_hint_from_xlsx = (
                                percent_hint_from_xlsx or cat_pct
                            )

                        values_raw_dec_xlsx: List[dict] = []
                        val_ref = ser.find(
                            ".//c:val//c:numRef//c:f", namespaces=NS
                        )
                        if val_ref is not None and val_ref.text:
                            val_vals, val_pct = read_range_plus(val_ref.text)
                            percent_hint_from_xlsx = (
                                percent_hint_from_xlsx or val_pct
                            )
                            for v in val_vals[:MAX_SERIES_POINTS]:
                                raw_str = "" if v is None else str(v)
                                d = _dec_from_any(v)
                                values_raw_dec_xlsx.append(
                                    {"raw": raw_str.strip(), "dec": d}
                                )

                        if cats and values_raw_dec_xlsx:
                            n = min(len(cats), len(values_raw_dec_xlsx))
                            cats = cats[:n]
                            values_raw_dec_xlsx = values_raw_dec_xlsx[:n]

                        series_tmp.append(
                            {
                                "name": series_name,
                                "cats": cats,
                                "values_raw_dec": values_raw_dec_xlsx,
                            }
                        )

                if series_tmp:
                    series = series_tmp

    # 3. Определяем, есть ли контекст процентов
    percent_by_fmt = (
        _has_percent_numfmt(root)
        or _is_percent_grouping(root)
        or percent_hint_from_xlsx
    )
    percent_by_values = False
    t = (chart_type or "").lower()

    if chart_type not in ("scatter", "bubble"):
        # Собираем данные для эвристики по значениям
        series_for_heur: List[Dict[str, List[float]]] = []
        for s in series:
            vals_f: List[float] = []
            for item in s.get("values_raw_dec", []):
                dec = item.get("dec")
                if dec is not None:
                    try:
                        vals_f.append(float(dec))
                    except Exception:
                        pass
            series_for_heur.append({"vals": vals_f})

        if t in ("pie", "ofpie", "doughnut"):
            percent_by_values = _sum_close_to_one_for_pie(series_for_heur)
        else:
            all_vals: List[float] = []
            for s in series_for_heur:
                all_vals.extend(s.get("vals", []))
            if all_vals and max(all_vals) <= 1.0:
                percent_by_values = _per_category_sums_close_to_one(
                    series_for_heur
                )

    percent_context = bool(percent_by_fmt or percent_by_values)

    # 4. Нормализуем данные серий
    if chart_type in ("scatter", "bubble"):
        # scatter / bubble: оставляем старую модель
        for s in series:
            s["x_vals"] = [
                float(v) for v in s.get("x_vals_num", []) if v != ""
            ]
            s["y_vals"] = [
                float(v) for v in s.get("y_vals_num", []) if v != ""
            ]
            if chart_type == "bubble":
                s["size_vals"] = [
                    float(v) for v in s.get("size_vals_num", []) if v != ""
                ]
    else:
        # категориальные: строим values / values_raw / unit + поддерживаем vals/vals_num
        for s in series:
            values_raw_dec = s.get("values_raw_dec", [])
            values: List[Optional[float]] = []
            values_raw: List[str] = []
            vals_num: List[str] = []

            # Если по формату/группировке понимаем, что это проценты –
            # просто помечаем unit='%', но не меняем сами числа.
            unit = "%" if percent_context else None

            for item in values_raw_dec:
                raw = (item.get("raw") or "").strip()
                dec = item.get("dec")
                values_raw.append(raw)

                if dec is None:
                    values.append(None)
                    vals_num.append("")
                    continue

                # БОЛЬШЕ НЕ домножаем на 100 — значение остаётся «как в документе»
                num = float(dec)
                values.append(num)

                # Строковое представление исходного числа (тоже «как есть»)
                s_dec = format(dec.normalize(), "f").rstrip("0").rstrip(".")
                vals_num.append(s_dec if s_dec != "" else "0")

            s["values"] = values
            s["values_raw"] = values_raw
            s["unit"] = unit
            s["vals_num"] = vals_num

            # Старое поле "vals" для совместимости: человекочитаемый вид,
            # но БЕЗ пересчитывания шкалы/нормировки.
            if percent_context:
                vals_str: List[str] = []
                for v in values:
                    if v is None:
                        vals_str.append("")
                        continue
                    if abs(v - round(v)) < 0.05:
                        sval = str(int(round(v)))
                    else:
                        sval = f"{v:.2f}".rstrip("0").rstrip(".")
                    vals_str.append(f"{sval}%")
                s["vals"] = vals_str
            else:
                s["vals"] = [v for v in values if v is not None]


    # 5. Общие категории (если совпадают)
    shared_cats: Optional[List[str]] = None
    if chart_type not in ("scatter", "bubble"):
        all_cats = [tuple(s.get("cats", [])) for s in series if s.get("cats")]
        if all_cats and all(x == all_cats[0] for x in all_cats):
            shared_cats = list(all_cats[0])
        # 5.1. Нормализованный список пар label/value/unit для простого использования
    # Теперь собираем ВСЕ серии, а не только первую.
    chart_data_rows: Optional[List[Dict[str, Any]]] = None
    if chart_type not in ("scatter", "bubble") and series:
        rows_all: List[Dict[str, Any]] = []
        for s in series:
            cats = s.get("cats") or shared_cats or []
            vals = s.get("values") or []
            unit = s.get("unit") or None
            sname = s.get("name")  # имя серии (Низкий/Средний/Высокий и т.п.)
            if not cats or not vals:
                continue
            n = min(len(cats), len(vals))
            for i in range(n):
                v = vals[i]
                if v is None:
                    # пропускаем только реально отсутствующие значения,
                    # 0.0 при этом сохраняем
                    continue
                rows_all.append(
                    {
                        "label": str(cats[i]).strip(),
                        "value": float(v),
                        "unit": unit,
                        "series_name": sname,
                    }
                )
        if rows_all:
            chart_data_rows = rows_all

    # 6. Флаг ok
    ok = False
    if chart_type in ("scatter", "bubble"):
        ok = bool(
            series
            and all(
                len(s.get("x_vals", [])) and len(s.get("y_vals", []))
                for s in series
            )
        )
    else:
        ok = bool(
            series
            and all(
                any(v is not None for v in s.get("values", [])) for s in series
            )
        )

    return {
        "type": chart_type or "unknown",
        "title": chart_title,
        "x_title": x_title,
        "y_title": y_title,
        "percent": percent_context,
        "cats": shared_cats,
        "series": series,
        "data": chart_data_rows,        # <-- простой вид
        "chart_data": chart_data_rows,  # <-- алиас специально для бота
        "ok": ok,
        "source": "ooxml",
        "provenance": {"chart_xml": chart_xml_path, "used": used_source},
    }

def _split_caption_num_tail(caption: Optional[str], *, is_table: bool = False) -> Tuple[Optional[str], Optional[str]]:
    """
    Из подписи вида 'Рисунок 2.3 — Структура системы'
    достаём:
      caption_num  -> '2.3' (нормализованный вид)
      caption_tail -> 'Структура системы'
    Для таблиц работает аналогично.
    """
    if not caption:
        return None, None
    cap = caption.strip()
    if not cap:
        return None, None

    m = (TAB_CAP_RE if is_table else FIG_CAP_RE).match(cap)
    if not m:
        return None, None

    num_raw = (m.group(1) or "").strip()
    num_norm = _norm_label_num(num_raw)
    # хвост — всё после номера, чистим разделители
    tail = cap[m.end(1):].lstrip(" .—–-:\u2013\u2014")
    if not tail:
        tail = None
    # caption_num теперь сразу в каноническом виде ("2.3")
    return num_norm or num_raw or None, tail


def build_index(docx_path: str) -> Dict:
    doc_id = uuid.uuid4().hex[:8]
    _ensure_dir(MEDIA_ROOT)
    _ensure_dir(INDEX_ROOT)
    with ZipFile(docx_path) as z:
        doc = _read_xml(z, "word/document.xml")
        docrels = _doc_rels(z, "word/_rels/document.xml.rels")
        headings: List[Heading] = []
        figures: List[Figure] = []
        tables: List[Table] = []
        fig_captions: List[Tuple[int, int, str]] = []
        tab_captions: List[Tuple[int, int, str]] = []

        content_seq: List[Tuple[str, int]] = []
        p_index = 0
        fig_idx = 0
        tbl_idx = 0

        for node in _iter_body_children(doc):
            if node.tag.endswith("p"):
                txt = _text_of(node)
                lvl = _extract_heading_level(node)
                if lvl and txt:
                    headings.append(Heading(level=lvl, text=txt))
                if txt:
                    m_fig = FIG_CAP_RE.match(txt)
                    if m_fig:
                        raw_num = (m_fig.group(1) or "").strip()
                        try:
                            # убираем литеры и пробелы в начале, берём целую часть до точки/запятой
                            core = re.sub(r"^[A-Za-zА-Яа-я]+\s*", "", raw_num)
                            main = re.split(r"[.,]", core)[0]
                            num_int = int(main) if main else None
                        except Exception:
                            num_int = None
                        if num_int is not None:
                            # запоминаем не только номер, но и индекс параграфа
                            fig_captions.append((p_index, num_int, txt))

                    m_tab = TAB_CAP_RE.match(txt)
                    if m_tab:
                        raw_num = (m_tab.group(1) or "").strip()
                        try:
                            core = re.sub(r"^[A-Za-zА-Яа-я]+\s*", "", raw_num)
                            main = re.split(r"[.,]", core)[0]
                            num_int = int(main) if main else None
                        except Exception:
                            num_int = None
                        if num_int is not None:
                            tab_captions.append((p_index, num_int, txt))


                drawings = node.findall(".//w:drawing", namespaces=NS)
                for d in drawings:
                    kind: Optional[str] = None
                    image_path: Optional[str] = None
                    chart_block: Optional[Dict[str, Any]] = None
                    gdata = d.find(".//a:graphic/a:graphicData", namespaces=NS)
                    if gdata is None:
                        continue
                    uri = gdata.get("uri", "")
                    if uri.endswith("/chart"):
                        cnode = gdata.find(".//c:chart", namespaces=NS)
                        if cnode is not None:
                            rid = cnode.get(f"{{{NS['r']}}}id")
                            if rid and rid in docrels:
                                tgt = docrels[rid]["Target"]
                                chart_xml = os.path.normpath(
                                    os.path.join("word", tgt)
                                ).replace("\\", "/")
                                if chart_xml in z.namelist():
                                    chart_block = _parse_chart_data(
                                        z, chart_xml
                                    )
                                    kind = "chart"
                    else:
                        blip = d.find(
                            ".//pic:pic/pic:blipFill/a:blip", namespaces=NS
                        )
                        if blip is None:
                            blip = d.find(".//a:blip", namespaces=NS)
                        rid = (
                            blip.get(f"{{{NS['r']}}}embed")
                            if blip is not None
                            else None
                        )
                        if rid and rid in docrels:
                            tgt = docrels[rid]["Target"]
                            image_path = _save_image_from_rel(z, tgt, doc_id)
                            if image_path:
                                kind = "image"
                    if kind:
                        figures.append(
                            Figure(
                                kind=kind,
                                image_path=image_path,
                                chart=chart_block,
                                _order=fig_idx,
                                para_idx=p_index,
                            )
                        )
                        content_seq.append(("figure", fig_idx))
                        fig_idx += 1
                content_seq.append(("p", p_index))
                p_index += 1
            elif node.tag.endswith("tbl"):
                rows = _parse_tbl(node, limit_rows=MAX_TABLE_ROWS)
                # привязываем таблицу к текущему p_index (последний параграф выше)
                tbl = Table(rows=rows, _order=tbl_idx, para_idx=p_index)
                tables.append(tbl)
                content_seq.append(("table", tbl_idx))
                tbl_idx += 1


        # пронумеруем фигуры/таблицы по подписям
                # пронумеруем фигуры/таблицы по подписям:
        # для каждой фигуры ищем ближайшую по индексу параграфа подпись "Рисунок N"
        used_fig_caps = set()
        for fig in figures:
            best = None
            best_dist = 10**9
            best_idx = None
            for idx, (cap_p_idx, cap_n, cap_txt) in enumerate(fig_captions):
                if idx in used_fig_caps:
                    continue
                if fig.para_idx >= 0:
                    dist = abs(cap_p_idx - fig.para_idx)
                else:
                    # фолбэк — старое поведение по порядку
                    dist = abs(idx - fig._order)
                if dist < best_dist:
                    best_dist = dist
                    best = (cap_n, cap_txt)
                    best_idx = idx
            # ограничиваемся небольшой "окрестностью" по тексту, чтобы не хватать чужие подписи
            if best is not None and best_dist <= 3 and best_idx is not None:
                cap_p_idx = fig_captions[best_idx][0]
                if cap_p_idx >= fig.para_idx:
                    fig.n, fig.caption = best
                    used_fig_caps.add(best_idx)
                else:
                    fig.n = (fig.n or (fig._order + 1))


        used_tab_caps = set()
        for tbl in tables:
            best = None
            best_dist = 10**9
            best_idx = None
            for idx, (cap_p_idx, cap_n, cap_txt) in enumerate(tab_captions):
                if idx in used_tab_caps:
                    continue
                if tbl.para_idx >= 0:
                    dist = abs(cap_p_idx - tbl.para_idx)
                else:
                    dist = abs(idx - tbl._order)
                if dist < best_dist:
                    best_dist = dist
                    best = (cap_n, cap_txt)
                    best_idx = idx
            if best is not None and best_dist <= 3:
                tbl.n, tbl.caption = best
                used_tab_caps.add(best_idx)
            else:
                tbl.n = (tbl.n or (tbl._order + 1))


        # список литературы
        references: List[str] = []
        paras = doc.findall(".//w:p", namespaces=NS)
        ref_mode = False
        for p in paras:
            txt = _text_of(p)
            if not txt:
                continue
            if not ref_mode and HEAD_REF_RE.match(txt):
                ref_mode = True
                continue
            if ref_mode:
                if _extract_heading_level(p):
                    break
                if txt.strip():
                    references.append(txt.strip())

        index: Dict[str, Any] = {
            "meta": {"doc_id": doc_id, "file": os.path.abspath(docx_path)},
            "headings": [{"level": h.level, "text": h.text} for h in headings],
            "figures": [],
            "tables": [],
            "references": references,
        }

        def _fig_sort_key(f: Figure):
            return (f.n if isinstance(f.n, int) else 10**9, f._order)

        for f in sorted(figures, key=_fig_sort_key):
            cap_num, cap_tail = _split_caption_num_tail(f.caption, is_table=False)
            # канонический номер, который будет использоваться дальше в БД/боте
            label_norm = _norm_label_num(cap_num if cap_num is not None else f.n)

            entry: Dict[str, Any] = {
                "n": f.n,                      # исходный int-номер
                "label": label_norm,           # нормализованный "2.3"
                "caption": f.caption,
                "caption_num": cap_num,        # строка из подписи (уже нормализована)
                "caption_tail": cap_tail,      # хвост подписи без "Рисунок 2.3 —"
                "kind": f.kind,
                "chart": f.chart if f.kind == "chart" else None,
                "image_path": f.image_path if f.kind == "image" else None,
            }

            if f.kind == "chart" and isinstance(entry["chart"], dict):
                # помечаем источник и поднимаем chart_data на верхний уровень фигуры
                entry["chart"]["source"] = "ooxml"
                chart_data = entry["chart"].get("chart_data") or entry["chart"].get("data")
                if chart_data:
                    entry["chart_data"] = chart_data

            index["figures"].append(entry)


        def _tbl_sort_key(t: Table):
            return (t.n if isinstance(t.n, int) else 10**9, t._order)

        for t in sorted(tables, key=_tbl_sort_key):
            cap_num, cap_tail = _split_caption_num_tail(t.caption, is_table=True)
            label_norm = _norm_label_num(cap_num if cap_num is not None else t.n)
            index["tables"].append(
                {
                    "n": t.n,
                    "label": label_norm,         # канонический "3.2"
                    "caption": t.caption,
                    "caption_num": cap_num,      # например "3.2"
                    "caption_tail": cap_tail,    # "Распределение ответов по ..."
                    "rows": t.rows,
                }
            )

        _ensure_dir(INDEX_ROOT)
        idx_path = os.path.join(INDEX_ROOT, f"{doc_id}.json")
        with open(idx_path, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False, indent=2, default=_json_default)
        return index



def _norm_label_num(s: Any) -> Optional[str]:
    """
    Нормализуем номер из подписи:
    'Рисунок А 2,3' -> '2.3'
    '2.3 '          -> '2.3'
    """
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    # убираем литеру в начале, пробелы и запятые
    s = re.sub(r"^[A-Za-zА-Яа-я]+\s*", "", s)
    s = s.replace("\u00A0", " ")
    s = s.replace(" ", "")
    s = s.replace(",", ".")
    return s or None


def _chart_rows_to_text(rows: List[Dict[str, Any]]) -> str:
    """
    Преобразуем chart_data (список {label, value, unit})
    в человекочитаемый текст с маркерами.

    Евристика:
    - если у строк unit содержит '%'
    - и ВСЕ числовые значения в диапазоне [0, 1.2],
      считаем, что это доли (0.8 → 80%) и домножаем на 100.
    """
    lines: List[str] = []

    # ---- Евристика для "долей" в процентах ----
    numeric_vals: List[float] = []
    has_percent_unit = False

    for r in rows or []:
        unit_raw = r.get("unit")
        if isinstance(unit_raw, str) and "%" in unit_raw:
            has_percent_unit = True

        val = r.get("value")
        if isinstance(val, Decimal):
            try:
                numeric_vals.append(float(val))
            except Exception:
                pass
        elif isinstance(val, (int, float)):
            try:
                numeric_vals.append(float(val))
            except Exception:
                pass

    is_share_like = bool(
        has_percent_unit
        and numeric_vals
        and all(0.0 <= v <= 1.2 for v in numeric_vals)
    )
    # -------------------------------------------

    for r in rows or []:
        label = str(r.get("label") or "").strip()
        series_name = str(r.get("series_name") or "").strip()
        val = r.get("value")
        unit = r.get("unit") or ""

        sval = ""
        v: Optional[float] = None

        if isinstance(val, Decimal):
            v = float(val)
        elif isinstance(val, (int, float)):
            v = float(val)
        else:
            # если вдруг value строка — просто берём как есть
            sval = str(val) if val is not None else ""

        # применяем евристику: 0.8 → 80 (если это доли в процентах)
        if v is not None and is_share_like:
            v = v * 100.0

        if sval == "" and v is not None:
            # форматируем число аккуратно
            if abs(v - round(v)) < 0.05:
                sval = str(int(round(v)))
            else:
                sval = f"{v:.2f}".rstrip("0").rstrip(".")

        if isinstance(unit, str) and unit.strip():
            # unit чаще всего '%'
            if "%" in unit and not sval.endswith("%"):
                sval = f"{sval}%"
            else:
                sval = f"{sval} {unit.strip()}"

        # если есть имя серии, добавляем его в скобках
        if series_name:
            label_full = f"{label} ({series_name})"
        else:
            label_full = label

        line = f"— {label_full}: {sval}".strip()
        if line:
            lines.append(line)

    return "\n".join(lines)


def figure_lookup(index: Dict, n: Any) -> Optional[Dict]:
    """
    Расширенный поиск фигуры:

    * принимает как int, так и строку "2.3", "Рисунок 2.3" и т.п.;
    * пытается сначала сопоставить caption_num ('2.3'),
      потом целую часть номера (n = 2);
    * если у фигуры есть chart_data, добавляет в результат
      поле 'values_text' — готовый текст со значениями диаграммы.
    """
    # нормализованный строковый номер вида "2.3"
    norm = _norm_label_num(n)

    # целая часть до точки как int (для совместимости со старым поведением)
    num_int: Optional[int] = None
    if norm is not None:
        try:
            num_int = int(norm.split(".")[0])
        except Exception:
            num_int = None
    else:
        try:
            num_int = int(n)  # старый путь
        except Exception:
            num_int = None

    figures = index.get("figures", []) or []

    # 1) пробуем совпадение по caption_num (полный номер "2.3")
    if norm is not None:
        for f in figures:
            cap = f.get("caption_num")
            if _norm_label_num(cap) == norm:
                res = dict(f)
                rows = (
                    res.get("chart_data")
                    or (res.get("chart") or {}).get("chart_data")
                    or (res.get("chart") or {}).get("data")
                )
                if isinstance(rows, list) and rows:
                    res["values_text"] = _chart_rows_to_text(rows)
                return res

    return None

def table_lookup(index: Dict, n: Any) -> Optional[Dict]:
    """
    Поиск таблицы по номеру.
    Понимает и int, и строки '3.2', 'Таблица 3.2' и т.п.
    Основной ключ — caption_num, фолбэк — целая часть n.
    """
    norm = _norm_label_num(n)

    num_int: Optional[int] = None
    if norm is not None:
        try:
            num_int = int(norm.split(".")[0])
        except Exception:
            num_int = None
    else:
        try:
            num_int = int(n)
        except Exception:
            num_int = None

    tables = index.get("tables", []) or []

    # 1) точное совпадение по caption_num ("3.2")
    if norm is not None:
        for t in tables:
            cap = t.get("caption_num")
            if _norm_label_num(cap) == norm:
                return t

    # 2) фолбэк по целой части
    if num_int is not None:
        for t in tables:
            if t.get("n") == num_int:
                return t

    return None


def purge_media(doc_id: str) -> None:
    p = os.path.join(MEDIA_ROOT, doc_id)
    if os.path.isdir(p):
        shutil.rmtree(p, ignore_errors=True)


def purge_index(doc_id: str) -> None:
    p = os.path.join(INDEX_ROOT, f"{doc_id}.json")
    if os.path.isfile(p):
        try:
            os.remove(p)
        except Exception:
            pass


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("usage: python -m app.ooxml_lite <path-to-docx>")
        raise SystemExit(1)
    idx = build_index(sys.argv[1])
    print(json.dumps(idx, ensure_ascii=False, indent=2, default=_json_default))
